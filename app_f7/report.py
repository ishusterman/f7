import openpyxl
from openpyxl.styles import Font, Border, Side, Color
from openpyxl.styles import Font, Fill
from .models import *
from .views import *


from django.http import HttpResponse

def date_convert_to_usa(date):
    date = date.split(".")
    res=date[2]+"-"+(date[1])+"-"+date[0]
    return res

# по queryset_res расчитываем вектор сумм
def get_sum(queryset_res):
    result = []
    num = queryset_res.aggregate(Sum('count'))
    result.append(num.get('count__sum'))

    num = queryset_res.aggregate(Sum('goin'))
    result.append(num.get('goin__sum'))

    num = queryset_res.aggregate(Sum('go_in_from_ds'))
    result.append(num.get('go_in_from_ds__sum'))

    num = queryset_res.aggregate(Sum('go_in_from_selo'))
    result.append(num.get('go_in_from_selo__sum'))

    num = queryset_res.aggregate(Sum('go_in_to_17'))
    result.append(num.get('go_in_to_17__sum'))

    num = queryset_res.aggregate(Sum('go_in_up_60'))
    result.append(num.get('go_in_up_60__sum'))

    num = queryset_res.aggregate(Sum('go_in_from_other_stac'))
    result.append(num.get('go_in_from_other_stac__sum'))

    num = queryset_res.aggregate(Sum('go_in_to_other_stac'))
    result.append(num.get('go_in_to_other_stac__sum'))

    num = queryset_res.aggregate(Sum('go_out'))
    result.append(num.get('go_out__sum'))

    num = queryset_res.aggregate(Sum('go_out_to_other_stac'))
    result.append(num.get('go_out_to_other_stac__sum'))

    num = queryset_res.aggregate(Sum('go_out_to_ds'))
    result.append(num.get('go_out_to_ds__sum'))

    num = queryset_res.aggregate(Sum('death'))
    result.append(num.get('death__sum'))

    num = queryset_res.aggregate(Sum('count_finish'))
    result.append(num.get('count_finish__sum'))

    num = queryset_res.aggregate(Sum('count_finish_selo'))
    result.append(num.get('count_finish_selo__sum'))

    num = queryset_res.aggregate(Sum('count_finish_mother'))
    result.append(num.get('count_finish_mother__sum'))

    return result

# отчет по отделения, раскадка по датам ОТ -- ДО
def create_report1(request):

    type_pay = request.POST['type_pay']
    if type_pay != "0":
        type_pay_name=s_type_pay.objects.get(id=type_pay).type_pay
    else:
        type_pay_name="ВСЕ"
    department = request.POST['department']
    department_name = s_department.objects.get(id=department).department
    date1 = request.POST['date1']
    date2 = request.POST['date2']
    date1_usa= date_convert_to_usa(date1)
    date2_usa = date_convert_to_usa(date2)

    queryset = t_move.objects.filter(date__gte=date1_usa).filter(date__lte=date2_usa).filter(department=department)
    if int(type_pay) > 0: # если не все виды оплаты
        queryset = queryset.filter(type_pay=type_pay) # тогда фильтр по виду оплаты

    wb = openpyxl.load_workbook(filename='report1.xlsx')
    sheet = wb['Лист1']

    i = 3 # стартовая строка

    date = datetime.strptime(date1_usa, '%Y-%m-%d')

    caption = "Отчет отделения " + department_name + " с " + date1 + " по " + date2 +". Вид оплаты: "+ type_pay_name
    sheet.cell(row=1, column=2, value=caption)

    while date <= datetime.strptime(date2_usa, '%Y-%m-%d'):

        i=i+1 # номер строки в экселе
        queryset_res = queryset.filter(date= date)
        vector = get_sum(queryset_res)
        f= sheet.cell(row=i, column=1, value=date.strftime("%Y-%m-%d"))
        f.border = Border(right=Side(style='thin'))
        j=2 # начинаем со второго столбца
        for item in vector:
            sheet.cell(row=i, column=j, value=item)
            j= j+1

        date = date + timedelta(days=1)  # следующий день

    # Итоги. БЕЗ ФИЛЬТРА по дате - складываем ВСЁ

    vector = get_sum(queryset)
    i = i + 1
    j=2
    f= sheet.cell(row=i, column=j-1, value="Итого")
    f.font = f.font.copy(bold=True, color = '00FF0000')
    f.border = Border(right=Side(style='thin'), top=Side(style='thin'))

    for item in vector:
        f= sheet.cell(row=i, column=j, value=item)
        f.font = f.font.copy(bold=True, color = '00FF0000')
        f.border = Border(top=Side(style='thin'))
        j = j + 1



    file_name = "report1.xlsx"

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+file_name
    wb.save(response)

    return response

# отчет по движению по датам ОТ -- ДО, реанимация отдельно
def create_report2(request):

    type_pay = request.POST['type_pay']
    if type_pay != "0":
        type_pay_name=s_type_pay.objects.get(id=type_pay).type_pay
    else:
        type_pay_name="ВСЕ"
    department = request.POST['department']
    department_name = s_department.objects.get(id=department).department
    date1 = request.POST['date1']
    date2 = request.POST['date2']
    date1_usa= date_convert_to_usa(date1)
    date2_usa = date_convert_to_usa(date2)

    queryset = t_move.objects.filter(date__gte=date1_usa).filter(date__lte=date2_usa)

    if int(type_pay) > 0: # если не все виды оплаты
        queryset = queryset.filter(type_pay=type_pay) # тогда фильтр по виду оплаты

    wb = openpyxl.load_workbook(filename='report1.xlsx')
    sheet = wb['Лист1']

    i = 3 # стартовая строка

    caption = "Отчет по движению с " + date1 + " по " + date2 +".Реанимация отдельно. Вид оплаты: "+ type_pay_name
    sheet.cell(row=1, column=2, value=caption)

    #department_table = s_department.objects.all()
    type_rean_table = s_type_rean.objects.all()
    profile_table = s_profile.objects.all()



    for profile_item  in profile_table:
        department_table = s_department.objects.filter(profile=profile_item.id)
        for department_item in department_table:
            i=i+1 # номер строки в экселе
            # по отделению по основному профилю, не реанимация
            print ("Отделение: " + department_item.department)
            queryset_res = queryset.filter(department=department_item.id, type_rean = 1)
            vector = get_sum(queryset_res)
            print (vector)
            f= sheet.cell(row=i, column=1, value=str(department_item))
            f.border = Border(right=Side(style='thin'))
            j=2 # начинаем со второго столбца
            for item in vector:
                sheet.cell(row=i, column=j, value=item)
                j= j+1

        # итог по профилю по основному, не реанимация
        i = i + 1
        queryset_res = queryset.filter(department__profile=profile_item.id, type_rean = 1)
        vector = get_sum(queryset_res)
        f= sheet.cell(row=i, column=1, value=str(profile_item))
        f.font = f.font.copy(bold=True)
        f.border = Border(bottom =Side(style='thin'),right= Side(style='thin'))
        j = 2  # начинаем со второго столбца
        for item in vector:
            f = sheet.cell(row=i, column=j, value=item)
            f.font = f.font.copy(bold=True)
            f.border = Border(bottom=Side(style='thin'))
            j = j + 1

    # Итоги по реанимации. БЕЗ ФИЛЬТРА по отделению  - складываем ВСЁ
    for type_rean_item in type_rean_table:
        if type_rean_item.id > 1: # если не основое отделение
            queryset_res = queryset.filter(type_rean=type_rean_item.id)
            vector = get_sum(queryset_res)
            i = i + 1
            j = 2
            f= sheet.cell(row=i, column=j - 1, value=str(type_rean_item))
            f.font = f.font.copy(bold=True)
            f.border = Border(right=Side(style='thin'))
            for item in vector:
                f= sheet.cell(row=i, column=j, value=item)
                f.font = f.font.copy(bold=True)
                j = j + 1


    # Окончательные итоги отчета. БЕЗ ФИЛЬТРА по отделению  - складываем ВСЁ

    vector = get_sum(queryset)
    i = i + 1
    j=2
    f=sheet.cell(row=i, column=j-1, value="Итого")
    f.font = f.font.copy(bold=True, color = '00FF0000')
    f.border = Border(bottom=Side(style='thin'), top=Side(style='thin'),right= Side(style='thin'))
    for item in vector:
        f= sheet.cell(row=i, column=j, value=item)
        f.font = f.font.copy(bold=True, color = '00FF0000')
        f.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))
        j = j + 1


    file_name = "report1.xlsx"

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+file_name
    wb.save(response)

    return response


# отчет по движению по датам ОТ -- ДО, подробно реанимация
def create_report3(request):

    type_pay = request.POST['type_pay']
    if type_pay != "0":
        type_pay_name=s_type_pay.objects.get(id=type_pay).type_pay
    else:
        type_pay_name="ВСЕ"
    department = request.POST['department']
    department_name = s_department.objects.get(id=department).department
    date1 = request.POST['date1']
    date2 = request.POST['date2']
    date1_usa= date_convert_to_usa(date1)
    date2_usa = date_convert_to_usa(date2)

    queryset = t_move.objects.filter(date__gte=date1_usa).filter(date__lte=date2_usa)

    if int(type_pay) > 0: # если не все виды оплаты
        queryset = queryset.filter(type_pay=type_pay) # тогда фильтр по виду оплаты

    wb = openpyxl.load_workbook(filename='report1.xlsx')
    sheet = wb['Лист1']

    i = 3 # стартовая строка

    caption = "Отчет по движению с " + date1 + " по " + date2 +".Подробно реанимация. Вид оплаты: "+ type_pay_name
    sheet.cell(row=1, column=2, value=caption)

    department_table = s_department.objects.all()
    type_rean_table = s_type_rean.objects.all()


    for department_item  in department_table:
        for type_rean_item in type_rean_table:
            i=i+1 # номер строки в экселе
            queryset_res = queryset.filter(department=department_item.id, type_rean=type_rean_item.id)
            vector = get_sum(queryset_res)
            f= sheet.cell(row=i, column=1, value=(str(department_item) + " " + str(type_rean_item)))
            f.border = Border(right=Side(style='thin'))

            j=2 # начинаем со второго столбца
            for item in vector:
                f= sheet.cell(row=i, column=j, value=item)

                j= j+1



    # Итоги. БЕЗ ФИЛЬТРА по отделению и типу реанимации - складываем ВСЁ

    vector = get_sum(queryset)
    i = i + 1
    j=2
    f= sheet.cell(row=i, column=j-1, value="Итого")
    f.font = f.font.copy(bold=True, color = '00FF0000')
    f.border = Border(right=Side(style='thin'), bottom=Side(style='thin'), top=Side(style='thin'))
    for item in vector:
        f=sheet.cell(row=i, column=j, value=item)
        f.font = f.font.copy(bold=True, color = '00FF0000')
        f.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        j = j + 1


    file_name = "report1.xlsx"

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+file_name
    wb.save(response)

    return response

# отчет по движению по датам ОТ -- ДО, реанимация отдельно не показываем
def create_report4(request):

    type_pay = request.POST['type_pay']
    if type_pay != "0":
        type_pay_name=s_type_pay.objects.get(id=type_pay).type_pay
    else:
        type_pay_name="ВСЕ"
    department = request.POST['department']
    department_name = s_department.objects.get(id=department).department
    date1 = request.POST['date1']
    date2 = request.POST['date2']
    date1_usa= date_convert_to_usa(date1)
    date2_usa = date_convert_to_usa(date2)

    queryset = t_move.objects.filter(date__gte=date1_usa).filter(date__lte=date2_usa)

    if int(type_pay) > 0: # если не все виды оплаты
        queryset = queryset.filter(type_pay=type_pay) # тогда фильтр по виду оплаты

    wb = openpyxl.load_workbook(filename='report1.xlsx')
    sheet = wb['Лист1']

    i = 3 # стартовая строка

    caption = "Отчет по движению с " + date1 + " по " + date2 +". Вид оплаты: "+ type_pay_name
    sheet.cell(row=1, column=2, value=caption)

    #department_table = s_department.objects.all()
    type_rean_table = s_type_rean.objects.all()
    profile_table = s_profile.objects.all()



    for profile_item  in profile_table:
        department_table = s_department.objects.filter(profile=profile_item.id)
        for department_item in department_table:
            i=i+1 # номер строки в экселе
            # по отделению по основному профилю, не реанимация
            print ("Отделение: " + department_item.department)
            queryset_res = queryset.filter(department=department_item.id)
            vector = get_sum(queryset_res)
            print (vector)
            f= sheet.cell(row=i, column=1, value=str(department_item))
            f.border = Border(right=Side(style='thin'))
            j=2 # начинаем со второго столбца
            for item in vector:
                sheet.cell(row=i, column=j, value=item)

                j= j+1

        # итог по профилю по всем типам реанимаций+основное
        i = i + 1
        queryset_res = queryset.filter(department__profile=profile_item.id)
        vector = get_sum(queryset_res)
        f= sheet.cell(row=i, column=1, value=str(profile_item))
        f.font = f.font.copy(bold=True)
        f.border = Border(right=Side(style='thin'),top=Side(style='thin'), bottom=Side(style='thin'))
        j = 2  # начинаем со второго столбца
        for item in vector:
            f= sheet.cell(row=i, column=j, value=item)
            f.font = f.font.copy(bold=True)
            f.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
            j = j + 1

    # Окончательные итоги отчета. БЕЗ ФИЛЬТРА по отделению  - складываем ВСЁ

    vector = get_sum(queryset)
    i = i + 1
    j=2
    f= sheet.cell(row=i, column=j-1, value="Итого")
    f.font = f.font.copy(bold=True, color = '00FF0000')
    f.border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for item in vector:
        f=sheet.cell(row=i, column=j, value=item)
        f.font = f.font.copy(bold=True, color = '00FF0000')
        f.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        j = j + 1


    file_name = "report1.xlsx"

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename='+file_name
    wb.save(response)

    return response